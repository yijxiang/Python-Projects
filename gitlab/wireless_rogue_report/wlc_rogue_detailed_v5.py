"""
v1 Uses speadsheet as data set
v2 Uses Prime API call as data set
v3 Uses WLC for Rogue list data set, lookup name by radio mac for highest_rssi_ap
v4 Add openpyxl and pandas to create .xlsx, can email to site
v5 Adapted for Gitlab and Vault
"""

import datetime, logging, netmiko, sys, time, smtplib, pandas, pynetbox, requests, argparse
from tqdm import tqdm
from functions import classify_rogue, site_email
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

parser=argparse.ArgumentParser(description="Email Wireless Rogue Spreadsheet")
parser.add_argument('--nb_token', required=True, help="Netbox Token")
parser.add_argument('--nb_url', required=True, help="Netbox URL")
parser.add_argument('--un', required=True, help="AD username")
parser.add_argument('--pw', required=True, help="AD password")

args = parser.parse_args()
un = args.un
pw = args.pw
nb_token = args.nb_token
nb_url = args.nb_url

logging.basicConfig(format='[%(levelname)s] %(asctime)s %(message)s', level=logging.INFO, 
    filename="wlc_rogue_detailed.log", datefmt='%m/%d/%Y - %H:%M:%S')
logger = logging.getLogger("wlc_backup")

requests.packages.urllib3.disable_warnings()
session = requests.Session()
session.verify = False

nb = pynetbox.api(url=nb_url, token=nb_token)
nb.http_session = session

try:
    # nb_wlcs = nb.dcim.devices.filter(name='WLCNAME')
    nb_wlcs = nb.dcim.devices.filter(role='wlc', tag='rogue_report', has_primary_ip=True, status='active')
    logger.info("Netbox WLC Collection Complete!")
except:
    logger.error("Failed to get WLCs from Netbox!")
    print("Failed to get WLCs from Netbox!")
    sys.exit()

print(f"\nCollecting Wireless Rogues For: ")
logger.info(f"Collecting Wireless Rogues For: ")
for wlc in nb_wlcs:
    print(wlc.name)
    logger.info(wlc.name)

created_files = []

for wlc in tqdm(nb_wlcs, desc="Collecting Rogue Details"):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Rogue SSIDs'
    data2 = []
    wlc_ip = wlc.primary_ip.address[:-3]
    wlc_name = wlc.name

    #Create Device Profile for Netmiko to Connect
    cisco_wlc = {
    'device_type': 'cisco_wlc_ssh', 
    'ip': wlc_ip, 
    'username': un, 
    'password': pw
    }

    site_code_name = wlc_name[3:-3] # strip first 3 and last 3 characters

    filetime1 = datetime.datetime.now().strftime("%m.%d.%Y") # File timestamp example 10.19.2020
    filename = wlc_name+"_Wireless_Rogues_"+filetime1+".xlsx"

    logger.info(f"Connecting to {wlc_name} - {wlc_ip}")
    print(f"Connecting to {wlc_name} - {wlc_ip}")
    try:
        net_connect = netmiko.ConnectHandler(**cisco_wlc)    # connect to the device w/ netmiko
    except:
        logger.error(f"Failed to connect to {wlc_name} - {wlc_ip}")
        logger.debug(f"Exception: {sys.exc_info()[0]}")

    wlc_r_sum = net_connect.send_command("show rogue ap summary", use_textfsm=True)
    logger.info("sending command - show rogue ap summary")
    time.sleep(1)

    wlc_advanced_sum_b = net_connect.send_command("show advanced 802.11b summary", use_textfsm=True)
    logger.info("sending command - show advanced 802.11b summary")
    time.sleep(1)

    wlc_advanced_sum_a = net_connect.send_command("show advanced 802.11a summary", use_textfsm=True)
    logger.info("sending command - show advanced 802.11a summary")
    time.sleep(1)

    for item in tqdm(wlc_r_sum, desc=f"{wlc_name} rogues"):
        mac_address = item['mac_address']
        rogue_clients = item['rogue_clients']
        det_aps = item['det_aps']
        highest_rssi = item['highest_rssi_det_ap']
        rssi1 = item['rssi_one']
        channel1 = item['channel_one']

        rogue_mac = mac_address  # Send command for each rogue in summary output
        this_cmd = net_connect.send_command("show rogue ap detailed "+f"{rogue_mac}", use_textfsm=True)
        logger.info(f"sending command - show rogue ap detailed {rogue_mac}")
        
        if this_cmd == "Rogue requested does not exist\n" or this_cmd == "\nRogue requested does not exist\n":
            logger.warning(f"Rogue {rogue_mac} no longer exists")
            continue

        else:
            try:
                ssid = this_cmd[1]['ssid']
            except:
                logger.warning(f"Currently no detailed info for {rogue_mac}.")
                continue
            last_seen = this_cmd[0]['last_seen']
            first_seen = this_cmd[0]['first_seen']
            classification = classify_rogue(ssid)
            for radio_b in wlc_advanced_sum_b:
                if radio_b['mac_address'] == highest_rssi:
                    highest_rssi = radio_b['ap_name']
            for radio_a in wlc_advanced_sum_a:
                if radio_a['mac_address'] == highest_rssi:
                    highest_rssi = radio_a['ap_name']

            row = {"MAC_Address": mac_address,"SSID": ssid, "Classification": classification, 
                "Rogue_Clients": rogue_clients, "Det-APs": det_aps, "Strongest_Det_AP": highest_rssi, 
                "Channel": channel1, "RSSI": rssi1, "Last_Seen": last_seen, "First_Seen": first_seen, }
            a = 0
            for item in this_cmd[1:]:
                a += 1
                row[f'Detecting_AP_{a}'] = item['reporting_ap_name']
                row[f'RSSI_{a}'] = item['rssi']

            data2.append(row)
    df = pandas.DataFrame(data=data2)
    for r in dataframe_to_rows(df, index=True, header=True):
        ws.append(r)

    for cell in ws['A'] + ws[1]:
        cell.style = 'Pandas'
    ws.delete_cols(idx=1)
    wb.save(f'./rogue_files/{filename}')

    ## FILE TO SEND AND ITS PATH
    SourcePathName  = './rogue_files/' + filename 

    cc_recipients = ['CC_EMAIL1', 'CC_EMAIL2']
    site_bisos = site_email(wlc_name) # UPDATE FUNCTION IF ADDING MORE WLCS

    msg = MIMEMultipart()
    msg['From'] = 'EMAIL'
    msg['To'] = site_bisos
    msg['Cc'] = ", ".join(cc_recipients)
    msg['Subject'] = f'{site_code_name} Rogue Wireless Report {filetime1}'
    text = ('Hello,\n\nPlease review the attached wireless rogue report and reply with any '
    'updates to the classification categories.\n\n'
    'Thank you,\n\n'
    'EMAIL SIGNATURE HERE')

    """
    html = ('Hello,<br><br>Please review the attached wireless rogue report and reply with any '
    'updates to the classification categories.<br><br>'
    'Thank you,<br><br>'
    'EMAIL SIGNATURE HERE')
    """

    part1 = MIMEText(text, 'plain')
    # part2 = MIMEText(html, 'html')

    msg.attach(part1) #Using Plain Text
    # msg.attach(part2)

    ## ATTACHMENT
    attachment = open(SourcePathName, 'rb')
    part = MIMEBase('application', "octet-stream")
    part.set_payload((attachment).read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f"attachment; filename={filename}")
    msg.attach(part)

    server = smtplib.SMTP("SMTP_SERVER")
    server.send_message(msg)
    msg = (f"To: {msg['To']} Cc: {msg['Cc']} with attachment {filename}")
    server.quit()
    created_files.append(msg)

print("\nCreated and Emailed Files:")
logger.info("Created and Emailed Files:")
for file in created_files:
    print(f"  {file}")
    logger.info(f"  {file}")
print("Finished\n")