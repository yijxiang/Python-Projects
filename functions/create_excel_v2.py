import sys
try:
    from openpyxl import Workbook
except:
    print("Error! openpyxl is not installed - try: \'pip install openpyxl\'")
    sys.exit()
from openpyxl.utils.dataframe import dataframe_to_rows
try:
    import pandas
except:
    print("Error! pandas is not installed - try: \'pip install pandas\'")
    sys.exit()

__author__ = "Louis Uylaki"

def create_excel_v2(datas, file_name, file_path, sheet_names):
    '''
    :Param datas - Build a list or lists of dictionaries
    :Param filename - Name of file without extention
    :Param file_path - Path to where the file will be saved i.e. /mnt/c/wsl_outputs/
    :Param sheet_names - List of Names of Sheet or Sheets in Workbook
    '''
    a = 0
    wb = Workbook()
    filename = f"{file_name}.xlsx"
    for data, name in zip(datas, sheet_names):
        a = a+1
        ws = f"ws{a}"

        if a == 1:
            ws = wb.active
            ws.title = name
        else:
            ws = wb.create_sheet(name)

        df = pandas.DataFrame(data=data)

        for r in dataframe_to_rows(df, index=True, header=True):
            ws.append(r)

        for cell in ws['A'] + ws[1]:
            cell.style = 'Pandas'
        ws.delete_cols(idx=1)
    try:
        wb.save(f'{file_path}{filename}')
        msg=f"\nFile \'{filename}\' saved to \'{file_path}\'.\nWill NOT overwrite existing files with same name.\n"
    except:
            msg=f"\nUnable to save file \'{filename}\' to location \'{file_path}\'.\n- Recommend confirming path and directory permissions.\n"
    return msg


"""
USAGE EXAMPLE:

from create_excel_v2 import create_excel_v2

data1=[
    {"test": "test_value_1"},
    {"test": "test_value_2"}
    ]

data2=[
    {"test2": "test_value_3"},
    {"test2": "test_value_4"}
    ]

datas = [data1,data2]
sheet_names = ['Sheet Name 1', 'Sheet Name 2']

file_name="Name_Your_File"
file_path='/path/to/file/'

new_spreadsheet=create_excel_v2(datas, file_name, file_path, sheet_names)
print(new_spreadsheet)

"""