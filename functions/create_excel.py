import sys
try:
    from openpyxl import Workbook
except:
    print("Error! openpyxl is not installed - try: \'pip install openpyxl\'")
    sys.exit()
from openpyxl.utils.dataframe import dataframe_to_rows
try:
    import pandas
except:
    print("Error! pandas is not installed - try: \'pip install pandas\'")
    sys.exit()

__author__ = "Louis Uylaki"

def create_excel(data, file_name, file_path, sheet_name):
    '''
    :Param data - Build a list of dictionaries
    :Param filename - Name of file without extention
    :Param file_path - Path to where the file will be saved i.e. /mnt/c/wsl_outputs/
    :Param sheet_name - Name of Sheet in Workbook
    '''

    wb = Workbook()
    ws = wb.active
    ws.title = f"{sheet_name}"
    filename = f"{file_name}.xlsx"

    df = pandas.DataFrame(data=data)

    for r in dataframe_to_rows(df, index=True, header=True):
        ws.append(r)

    for cell in ws['A'] + ws[1]:
        cell.style = 'Pandas'
    ws.delete_cols(idx=1)
    try:
        wb.save(f'{file_path}{filename}')
        msg=f"\nFile \'{filename}\' saved to \'{file_path}\'.\nWill NOT overwrite existing files with same name.\n"
    except:
        msg=f"\nUnable to save file \'{filename}\' to location \'{file_path}\'.\n- Recommend confirming path and directory permissions.\n"
    return msg

"""
USAGE EXAMPLE:

from create_excel import create_excel

data=[
    {"test": "test_value_1"},
    {"test": "test_value_2"}
    ]
file_name="Name_Your_File"
file_path='/path/to/file/'
sheet_name='My Sheet Name'

new_spreadsheet=create_excel(data, file_name, file_path, sheet_name)
print(new_spreadsheet)

"""