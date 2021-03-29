import sys
try:
    from openpyxl import Workbook
except:
    print("Error openpyxl is not installed - try: \'pip install openpyxl\'")
    sys.exit()
from openpyxl.utils.dataframe import dataframe_to_rows
try:
    import pandas
except:
    print("Error pandas is not installed - try: \'pip install pandas\'")
    sys.exit()

__author__ = "Louis Uylaki"

class Spreadsheet:
    def __init__(self, data, file_name, file_path, sheet_name ):
        self.file_name=file_name
        self.file_path=file_path
        self.sheet_name=sheet_name
        self.data=data

    def create_excel(self):
        '''
        :Param data - Build a list of dictionaries
        :Param filename - Name of file without extention
        :Param file_path - Path to where the file will be saved i.e. /mnt/c/wsl_outputs/
        :Param sheet_name - Name of Sheet in Workbook
        '''

        wb = Workbook()
        ws = wb.active
        ws.title = f"{self.sheet_name}"
        filename = f"{self.file_name}.xlsx"

        df = pandas.DataFrame(data=self.data)

        for r in dataframe_to_rows(df, index=True, header=True):
            ws.append(r)

        for cell in ws['A'] + ws[1]:
            cell.style = 'Pandas'
        ws.delete_cols(idx=1)
        try:
            wb.save(f'{self.file_path}{filename}')
            msg=f"\nFile \'{filename}\' saved to \'{self.file_path}\'.\nWill NOT overwrite existing files with same name.\n"
        except:
            msg=f"\nUnable to save file \'{filename}\' to location \'{self.file_path}\'.\n- Recommend confirming path and directory permissions.\n"
        return msg

"""
USAGE EXAMPLE:

from create_excel_class import Spreadsheet

data=[
    {"test": "test_value_1"},
    {"test": "test_value_2"}
    ]
file_name="Testing3"
file_path='/mnt/c/wsl_outputs/'
sheet_name='My Sheet Name'

new_spreadsheet=Spreadsheet(data, file_name, file_path, sheet_name)
a=new_spreadsheet.create_excel()

print(a)

"""