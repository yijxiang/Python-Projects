
import sys, requests, pynetbox, netmiko, datetime, logging, pandas, hvac
from secrets import un, pw
from tqdm import tqdm
from rich import print
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

logging.basicConfig(format='[%(levelname)s] %(asctime)s %(message)s', level=logging.INFO, 
filename = "wlc_dca.log", datefmt='%m/%d/%Y - %H:%M:%S')
logger = logging.getLogger("wlc_backup")

requests.packages.urllib3.disable_warnings()
session = requests.Session()
session.verify = False

wb = Workbook()
ws = wb.active

def retrieve_secret(un, pw, mount_point, path):
    client = hvac.Client(url='https://VAULT:PORT', verify=False)
    login_response = client.auth.radius.login(
        un, pw,
        use_token = True,
        mount_point = 'radius/config'
    )
    token = login_response['auth']['client_token']
    client = hvac.Client(url='https://VAULT:PORT', token=token, verify=False)
    read_response = client.secrets.kv.read_secret_version(mount_point=mount_point, path=path)
    return read_response['data']['data']

nb_secrets = retrieve_secret(un, pw, un, 'netbox')
nb_url = nb_secrets['nb_url']
nb_token = nb_secrets['nb_token']

nb = pynetbox.api(url=nb_url, token=nb_token)
nb.http_session = session

try:
    # nb_wlcs = nb.dcim.devices.filter(name='WLCNAME')
    # nb_wlcs = nb.dcim.devices.filter(role='wlc-lab', has_primary_ip=True, status='active')
    nb_wlcs = nb.dcim.devices.filter(role='wlc', tag='wlc_backups', has_primary_ip=True, status='active')
    logger.info("Netbox WLC Collection Complete!")
except:
    logger.error("Failed to get WLCs from Netbox!")
    print("Failed to get WLCs from Netbox!")
    sys.exit()

print(f"\nCollecting 5GHz DCA Channels for: ")
logger.info(f"Collecting 5GHz DCA Channels for: ")
for wlc in nb_wlcs:
    print(wlc.name)
    logger.info(wlc.name)
print("\n")

ssh_errors = []
channel_list = []
for wlc in tqdm(nb_wlcs, desc="Collecting DCA Channels"):
    wlc_ip = wlc.primary_ip.address[:-3]
    wlc_name = wlc.name

    #Create Device Profile for Netmiko to Connect
    cisco_wlc = {
    'device_type': 'cisco_wlc_ssh', 
    'ip': wlc_ip, 
    'username': un, 
    'password': pw
    }

    try:
        net_connect = netmiko.ConnectHandler(**cisco_wlc)    # connect to the device w/ netmiko
    except:
        ssh_errors.append(f"[red]  ERROR! Failed to connect to {wlc_name} - {cisco_wlc['ip']}[/red]")
        logger.error(f"Failed to connect to {wlc_name} - {cisco_wlc['ip']}")
        logger.debug(f"Exception: {sys.exc_info()[0]}")
        continue

    show_channels = net_connect.send_command("show advanced 802.11a channel", use_textfsm=True)
    five_ghz_dca_channels = show_channels[0]['dca_channels']
    new_list = []
    
    for ch in five_ghz_dca_channels:
        split_ch = ch.split(',')
        for item in split_ch:
            if item is not '':
                new_list.append(item)
    list_string = ', '.join(new_list)
    data = {"WLC_NAME": wlc_name, "Channel_Width": show_channels[0]['dca_width'], "Update_Interval": show_channels[0]['update_time'], 
        "DCA_Sensitivity_Level": show_channels[0]['dca_sensitivity'], "Assign_Mode": show_channels[0]['assign_mode'], "Used_DCA_Channels": list_string}
    logger.info(data)
    channel_list.append(data)

filetime1 = datetime.datetime.now().strftime("%m.%d.%y-%I.%M%p") # File timestamp example 12.21.20-09.53AM
filename = f"WLC_5GHz_DCA_Channels_{filetime1}.xlsx"
df = pandas.DataFrame(data=channel_list)

for r in dataframe_to_rows(df, index=True, header=True):
    ws.append(r)

for cell in ws['A'] + ws[1]:
    cell.style = 'Pandas'
ws.delete_cols(idx=1)
wb.save(f'/mnt/c/wsl_outputs/wlc_dca_channels/{filename}')

for error in ssh_errors:
        print(f"[red bold]{error}[/red bold]")
print("\n")
print(f"Created file C:\wsl_outputs\wlc_dca_channels\{filename}\n")