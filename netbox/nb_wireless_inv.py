'''
Will Create Spreadsheet Of All Active Wireless Equipment.
Exludes Lab
'''
import pynetbox, requests, datetime, pandas
from secrets import nb_token, nb_url
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

session = requests.Session()
session.verify = False
requests.packages.urllib3.disable_warnings() 

nb = pynetbox.api(url=nb_url, token=nb_token) # Define Nebox API Call
nb.http_session = session

# devices=nb.dcim.devices.filter(site="emp", role=["access", "access-point", "distribution"], has_primary_ip=True)     #Get all access devices in Netbox from Corp9
devices = nb.dcim.devices.filter(role=["wlc", "access-point", "bridge"], status="active")
sites = nb.dcim.sites.filter(region=["canada", "mexico", "us"])
files = [] # Empty list to store filenames
data2 = [] # Stores data dictionaries

filetime = datetime.datetime.now().strftime("%m.%d.%y-%H.%M.%S") # File timestamp
config_filename = "Wireless-Equipment-List" + "_" + filetime + ".xlsx" # Create File Name, Show_Inventory with timestamp and extention
files.append(config_filename)    # Append File Name to List for Output and End of Script

for device in devices:
    if device.primary_ip is None:
        ip = "N/A"
    else:
        ip = device.primary_ip.address[:-3]
    name = device.name
    serial = device.serial
    model = device.device_type.model
    role = device.device_role.name
    manufacturer = device.device_type.manufacturer.name
    site = device.site.name

    for a in sites:
        if a.name == site:
            facility = a.facility
            address = a.physical_address
            region = a.region.name

    data = {"Site_ID": site, "Name": name, "Device_Role": role, "Manufacturer": manufacturer, "Model": model, "Serial": serial, 
        "IP_Address": ip, "Region": region, "Facility": facility, "Address": address}
    data2.append(data)

wb = Workbook()
ws = wb.active

df = pandas.DataFrame(data=data2)
for r in dataframe_to_rows(df, index=True, header=True):
    ws.append(r)
for cell in ws['A'] + ws[1]:
    cell.style = 'Pandas'
ws.delete_cols(idx=1)
wb.save(f'/mnt/c/wsl_outputs/inventory/{config_filename}') 

print("Finished Creating Files:")
for fname in files:
    print(f"./files/{fname}")